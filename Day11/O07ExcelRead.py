
from openpyxl import load_workbook

wb = load_workbook("FirstExcel.xlsx")

ws = wb.active

print(ws.dimensions)

dataRange = ws[ws.dimensions]

print(dataRange)

print("#" * 40)
print("\n")
for c1, c2, c3, c4, c5 in dataRange:
    print("{0:20}\t\t{1:^5}\t\t{2:5}\t\t{3:5}\t\t{4:5}".format(c1.value,
                c2.value, c3.value, c4.value, c5.value))
