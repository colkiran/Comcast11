
from openpyxl import load_workbook

wb = load_workbook("FirstExcel.xlsx")

wb.create_sheet("Players")

wb.active = wb['Players']

wb['Comcast'].active = False

wb.active = wb['Players']

ws = wb.active

cell = ws["A5"]

data  = [
    ['PlayerName', "Age", "Runs", "Matches", "Country"],
    ['Sachin Tendulkar', 49, 38650, 845, 'India'],
    ['Brain Lara', 52, 26780, 670, "West Indies"],
    ['Ricky Ponting', 46, 29750, 630, "Australia"],
    ['Jayasurya', 48, 22300, 580, "Sri Lanka"],
    ['Inzamam ul haq', 45, 26500, 590, "Pakistan"]
]

for row in data:
    ws.append(row)

wb.save("FirstExcel.xlsx")
