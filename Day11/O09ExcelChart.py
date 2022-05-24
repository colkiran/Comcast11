
from openpyxl import load_workbook
from openpyxl.chart import  Reference, BarChart, Series

wb = load_workbook("FirstExcel.xlsx")

wb.create_sheet("Chart")

wb.active = wb['Chart']

ws = wb.active

prod = [
    ['Products', 'Sales'],
    ['pepsi', 350],
    ['coke', 235],
    ['thumbs up', 405],
    ['fanta', 185],
    ['sprite', 250],
    ['maaza', 300],
    ['slice', 200],
    ['mirinda', 280]
]

for row in prod:
    ws.append(row)

dataRef = Reference(ws, min_row=2, min_col=2, max_row=9)
labelRef = Reference(ws, min_row=2, min_col=1, max_row=9)

chart = BarChart()
chart.add_data(dataRef)
chart.set_categories(labelRef)
# series = Series(values=dataRef)
# chart.series.append(series)
chart.x_axis.title = "Products"
chart.y_axis.title = "Sales in lakhs"
chart.title = "Baverage Sales"

ws.add_chart(chart, "E9")
wb.save("FirstExcel.xlsx")
