
from openpyxl import load_workbook

wb = load_workbook("FirstExcel.xlsx")

ws = wb.active

print(ws.dimensions)
# A6 : E11  => rows = 6 - 11; col = 1 - 5

for row in ws.iter_rows(min_row=6, min_col=1, max_row=11, max_col=5):
    for cell in row:
        if cell.value == "Jayasurya":
            cell.value = "sanath jayasurya".upper()

wb.save("FirstExcel.xlsx")