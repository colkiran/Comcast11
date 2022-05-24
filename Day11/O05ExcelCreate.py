
"""
Excel file
----------
file  - workbook
pages - worksheets

worksheet => grid like structure
grids will have

1. rows - numbered between 1 and 1048576 rows
2. cols - will be labeled like A - Z, AA - AZ, BA - BZ.....XFD (16384 cols)

data will be entered into a cell which is an intersection of row and column
each is addressed with a combination of colname and row number
for eg: col- B and row - 10 => we call it as B10


"""


from openpyxl import Workbook
from datetime import datetime

wb = Workbook()

ws = wb.active

ws.title = "Comcast"

ws["B5"] = "Hello World"
ws["C5"] = 3500
ws['D5'] = datetime.now()

wb.save("FirstExcel.xlsx")